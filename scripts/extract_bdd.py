"""
Script d'extraction BDD Excel → JSON
Source : Calculateur de performances thermiques JANNEAU.xlsm (v1.6K)

Exécution :
    python scripts/extract_bdd.py

Génère dans backend/app/data/ :
    - chassis.json        : liste des types de châssis
    - vitrages.json       : base des compositions de vitrage
    - psi_g.json          : coefficients linéiques d'intercalaire (Psi g)
    - couleurs.json       : couleurs + coefficient d'absorption α
    - volets.json         : volets roulants (rR, Uc coefficients)
"""

import json
import re
from pathlib import Path

import openpyxl

EXCEL_FILE = Path(__file__).parent.parent / "Calculateur de performances thermiques JANNEAU.xlsm"
OUTPUT_DIR = Path(__file__).parent.parent / "backend" / "app" / "data"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

print(f"Lecture de : {EXCEL_FILE}")
wb = openpyxl.load_workbook(EXCEL_FILE, read_only=True, keep_vba=True, data_only=True)


# ---------------------------------------------------------------------------
# 1. CHASSIS — onglet BDD lignes 13 à ~330
# ---------------------------------------------------------------------------
def extract_chassis(ws) -> list[dict]:
    """
    Colonnes BDD (0-indexé) :
      B(1) = nom du châssis
      C(2) = id_onglet (numéro de l'onglet de calcul)
      D(3) = nb_vitrages
      E(4) = nb_panneaux
    Les lignes de catégorie (id=0) sont ignorées.
    """
    chassis_list = []
    current_famille = None
    current_gamme = None

    FAMILLE_MARKERS = {
        "FENETRES A FRAPPE": "FENETRES A FRAPPE",
        "PORTES CREMONE-SERRURE-SERVICE": "PORTES CREMONE-SERRURE-SERVICE",
        "BAIES COULISSANTES": "BAIES COULISSANTES",
        "PORTES D'ENTREE": "PORTES D'ENTREE",
    }

    # Mapping code gamme interne (utilisé pour Psi_g et couleurs)
    GAMME_CODE_MAP = {
        "BOIS": "BOIS",
        "CARLIS.J": "CARL",
        "LITTORAL.J": "CARL",
        "PRIMELIS": "PRIM",
        "ANTALIS": "PRIM",
        "EVOLUTION": "EVOL",
        "ESSENTIEL": "ESSE",
        "SOLARIS II": "SOII",
        "SOLARIS": "SOLA",
        "CG-ALU": "CG-A",
        "CAPITALES": "INNO",
        "CAPITALES MD76": "INNO",
        "FLEUVES ET RIVIERES": "INNO",
        "HYLLIADE": "HYLL",
        "LUMIS - BOIS": "BOIS",
        "LUMIS - PVC": "INNO",
        "LUMIS - ALU": "ALLI",
        "ALLIS": "ALLI",
    }

    for row in ws.iter_rows(min_row=13, max_row=330, values_only=True):
        nom = row[1]
        id_onglet = row[2]
        nb_vitrages = row[3]
        nb_panneaux = row[4]

        if nom is None:
            continue

        nom = str(nom).strip()

        # Détection des familles principales
        for marker, famille in FAMILLE_MARKERS.items():
            if nom == marker:
                current_famille = famille
                break

        # Lignes de sous-catégorie (id=0 ou non numérique)
        if not isinstance(id_onglet, (int, float)) or id_onglet == 0:
            # Peut être un marqueur de gamme (ligne indentée avec des espaces)
            clean = nom.lstrip()
            if clean and nom.startswith("          "):
                current_gamme = clean
            continue

        id_onglet = int(id_onglet)
        nb_vitrages = int(nb_vitrages) if isinstance(nb_vitrages, (int, float)) else 0
        nb_panneaux = int(nb_panneaux) if isinstance(nb_panneaux, (int, float)) else 0

        # Déterminer le code gamme interne à partir du nom du châssis
        gamme_code = "INNO"  # défaut
        for prefix, code in GAMME_CODE_MAP.items():
            if nom.upper().startswith(prefix.upper()):
                gamme_code = code
                break
        # Cas BOIS
        if nom.upper().startswith("BOIS"):
            gamme_code = "BOIS"

        chassis_list.append({
            "id": id_onglet,
            "nom": nom,
            "famille": current_famille,
            "gamme": current_gamme,
            "gamme_code": gamme_code,
            "nb_vitrages": nb_vitrages,
            "nb_panneaux": nb_panneaux,
        })

    return chassis_list


# ---------------------------------------------------------------------------
# 2. VITRAGES — onglet BDD lignes 527 à ~1076
# ---------------------------------------------------------------------------
def extract_vitrages(ws) -> list[dict]:
    """
    Colonnes (0-indexé) :
      B(1) = composition
      C(2) = Ug
      D(3) = Sg
      E(4) = Sg1
      F(5) = Sg2
      G(6) = Sg3
      H(7) = Tlg
    """
    vitrages = []
    current_type = "DOUBLE VITRAGE"

    for row in ws.iter_rows(min_row=527, max_row=1076, values_only=True):
        composition = row[1]
        ug = row[2]

        if composition is None:
            continue
        composition = str(composition).strip()

        # Marqueurs de section
        if composition in ("DOUBLE VITRAGE", "TRIPLE VITRAGE", "PANNEAUX"):
            current_type = composition
            continue

        # Ignorer les en-têtes
        if composition in ("Vitrage (int/ext)", "Ug", "Composition"):
            continue

        if not isinstance(ug, (int, float)):
            continue

        sg = row[3] if isinstance(row[3], (int, float)) else None
        sg1 = row[4] if isinstance(row[4], (int, float)) else None
        sg2 = row[5] if isinstance(row[5], (int, float)) else None
        sg3 = row[6] if isinstance(row[6], (int, float)) else None
        tlg = row[7] if isinstance(row[7], (int, float)) else None

        vitrages.append({
            "composition": composition,
            "type": current_type,
            "Ug": round(float(ug), 4),
            "Sg": round(float(sg), 4) if sg is not None else None,
            "Sg1": round(float(sg1), 4) if sg1 is not None else None,
            "Sg2": round(float(sg2), 4) if sg2 is not None else None,
            "Sg3": round(float(sg3), 4) if sg3 is not None else None,
            "Tlg": round(float(tlg), 4) if tlg is not None else None,
        })

    return vitrages


# ---------------------------------------------------------------------------
# 3. PSI_G — onglet BDD lignes 1349 à 1766
# ---------------------------------------------------------------------------
def extract_psi_g(ws) -> list[dict]:
    """
    Colonnes (0-indexé) :
      B(1) = gamme (BOIS, INNO, CARL, EXCL, SOLA, SOII, HYLL, ALLI, CG-A...)
      C(2) = intercalaire (Aluminium, Warm-Edge, Swisspacer...)
      E(4) = renfort (Non / Oui)
      F(5) = Ug
      G(6) = Psi_g
    Clé VLOOKUP Excel : CONCATENATE(gamme_code, intercalaire, Ug_fr)
    On stocke les 4 champs pour une recherche exacte côté Python.
    """
    psi_g_list = []

    for row in ws.iter_rows(min_row=1349, max_row=1766, values_only=True):
        gamme = row[1]
        intercalaire = row[2]
        renfort = row[4]
        ug = row[5]
        psi_g = row[6]

        if gamme is None or intercalaire is None or psi_g is None:
            continue
        if not isinstance(psi_g, (int, float)):
            continue

        # Normaliser Ug (peut être stocké comme string "0.7" dans quelques cas)
        try:
            ug_float = float(str(ug).replace(",", "."))
        except (ValueError, TypeError):
            continue

        psi_g_list.append({
            "gamme": str(gamme).strip(),
            "intercalaire": str(intercalaire).strip(),
            "renfort": str(renfort).strip() if renfort else "Non",
            "Ug": round(ug_float, 4),
            "Psi_g": round(float(psi_g), 5),
        })

    return psi_g_list


# ---------------------------------------------------------------------------
# 4. COULEURS — onglet BDD lignes ~259 à ~441
# ---------------------------------------------------------------------------
def extract_couleurs(ws) -> list[dict]:
    """
    Les couleurs sont organisées par famille (Famille N°1, N°2, etc.)
    Colonnes (0-indexé) :
      B(1) = nom couleur
      C(2) = code α (1=Claire, 2=Moyenne, 3=Sombre, 4=Noire)
    """
    ALPHA_MAP = {1: 0.4, 2: 0.6, 3: 0.8, 4: 1.0}
    couleurs = []
    current_famille = None

    for row in ws.iter_rows(min_row=259, max_row=441, values_only=True):
        cell_a = row[0]
        nom = row[1]
        code = row[2]

        # Détection famille
        if cell_a and str(cell_a).startswith("Famille N°"):
            current_famille = str(cell_a).strip()
            continue

        if nom is None or not isinstance(code, (int, float)):
            continue

        nom = str(nom).strip()
        if not nom or nom.startswith("Coefficient") or nom.startswith("Code"):
            continue

        code_int = int(code)
        alpha = ALPHA_MAP.get(code_int, 0.6)

        couleurs.append({
            "nom": nom,
            "famille_prix": current_famille,
            "code_alpha": code_int,
            "alpha": alpha,
        })

    return couleurs


# ---------------------------------------------------------------------------
# 5. VOLETS — onglet BDD lignes ~443 à ~530
# ---------------------------------------------------------------------------
def extract_volets(ws) -> list[dict]:
    """
    Section volets : deux tableaux
    A) Tableau principal (lignes ~448-475) :
       B(1) = désignation volet
       C-H  = compatibilité gammes (X si compatible)
       I(8) = rR [m²K/W]
       J(9) = hauteur_coffre [m]

    B) Tableau Uc = a + b/L (lignes ~480-520) :
       B(1) = désignation volet
       C(2) = a_T,  D(3) = b_T
       E(4) = a_P0, F(5) = b_P0
       G(6) = a_P1, H(7) = b_P1
       I(8) = a_P2, J(9) = b_P2
       K(10)= a_P3, L(11)= b_P3
       M(12)= a_P4, N(13)= b_P4
    """
    GAMME_COLS = {2: "BOIS", 3: "INNO", 4: "CARL", 5: "EXCL", 6: "SOLA", 7: "SOII"}
    ISOLATION_LEVELS = ["T", "P0", "P1", "P2", "P3", "P4"]

    # --- Tableau principal rR ---
    volets_base = {}
    for row in ws.iter_rows(min_row=448, max_row=476, values_only=True):
        nom = row[1]
        if nom is None or not isinstance(nom, str):
            continue
        nom = nom.strip()
        if not nom or nom.startswith("Avez") or nom.startswith("Quel"):
            continue
        rR = row[8]
        hauteur = row[9]
        if not isinstance(rR, (int, float)):
            continue

        gammes_compatibles = []
        for col_idx, gamme in GAMME_COLS.items():
            val = row[col_idx]
            if val == "X" or val == "x":
                gammes_compatibles.append(gamme)

        volets_base[nom] = {
            "designation": nom,
            "gammes_compatibles": gammes_compatibles,
            "rR": round(float(rR), 4),
            "hauteur_coffre": round(float(hauteur), 4) if isinstance(hauteur, (int, float)) else None,
            "uc_coefficients": {},
        }

    # --- Tableau Uc coefficients ---
    for row in ws.iter_rows(min_row=480, max_row=520, values_only=True):
        nom = row[1]
        if nom is None or not isinstance(nom, str):
            continue
        nom = nom.strip()
        if nom not in volets_base:
            continue

        uc = {}
        col_pairs = [
            ("T", 2, 3), ("P0", 4, 5), ("P1", 6, 7),
            ("P2", 8, 9), ("P3", 10, 11), ("P4", 12, 13),
        ]
        for level, a_col, b_col in col_pairs:
            a = row[a_col]
            b = row[b_col]
            if isinstance(a, (int, float)) and isinstance(b, (int, float)):
                uc[level] = {"a": round(float(a), 5), "b": round(float(b), 5)}

        if uc:
            volets_base[nom]["uc_coefficients"] = uc

    return list(volets_base.values())


# ---------------------------------------------------------------------------
# 6. INTERCALAIRES — liste unique extraite de psi_g
# ---------------------------------------------------------------------------
def extract_intercalaires(psi_g_list: list[dict]) -> list[str]:
    seen = set()
    result = []
    for entry in psi_g_list:
        name = entry["intercalaire"]
        if name not in seen:
            seen.add(name)
            result.append(name)
    return sorted(result)


# ---------------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------------
def main():
    ws_bdd = wb["BDD"]

    print("Extraction des châssis...")
    chassis = extract_chassis(ws_bdd)
    print(f"  → {len(chassis)} châssis extraits")

    print("Extraction des vitrages...")
    vitrages = extract_vitrages(ws_bdd)
    print(f"  → {len(vitrages)} compositions de vitrage extraites")

    print("Extraction des Psi_g (intercalaires)...")
    psi_g = extract_psi_g(ws_bdd)
    intercalaires = extract_intercalaires(psi_g)
    print(f"  → {len(psi_g)} entrées Psi_g, {len(intercalaires)} types d'intercalaire")

    print("Extraction des couleurs...")
    couleurs = extract_couleurs(ws_bdd)
    print(f"  → {len(couleurs)} couleurs extraites")

    print("Extraction des volets...")
    volets = extract_volets(ws_bdd)
    print(f"  → {len(volets)} types de volets extraits")

    # Écriture des fichiers JSON
    files = {
        "chassis.json": chassis,
        "vitrages.json": vitrages,
        "psi_g.json": psi_g,
        "intercalaires.json": intercalaires,
        "couleurs.json": couleurs,
        "volets.json": volets,
    }

    for filename, data in files.items():
        path = OUTPUT_DIR / filename
        path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
        print(f"  Écrit : {path}")

    print("\nExtraction BDD terminée.")


if __name__ == "__main__":
    main()
