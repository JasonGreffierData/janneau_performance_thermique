"""
Script d'extraction de la géométrie des châssis Excel → JSON
Source : onglets 1 à 266 du fichier JANNEAU.xlsm

Stratégie :
  - data_only=True : lit les valeurs Af et Uf calculées par les VLOOKUP
  - data_only=False : lit les formules pour identifier le type géométrique
  - Le type géométrique est déduit des noms de pièces (H column)

Exécution :
    python scripts/extract_geometry.py

Génère : backend/app/data/chassis_geometry.json
"""

import json
import re
from pathlib import Path

import openpyxl

EXCEL_FILE = Path(__file__).parent.parent / "Calculateur de performances thermiques JANNEAU.xlsm"
OUTPUT_DIR = Path(__file__).parent.parent / "backend" / "app" / "data"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

# Charger les deux versions : valeurs (data_only) et formules (pas data_only)
print("Chargement du fichier Excel (valeurs)...")
wb_val = openpyxl.load_workbook(EXCEL_FILE, read_only=True, keep_vba=True, data_only=True)
print("Chargement du fichier Excel (formules)...")
wb_fml = openpyxl.load_workbook(EXCEL_FILE, read_only=True, keep_vba=True, data_only=False)

# Liste des chassis depuis chassis.json
chassis_list = json.loads((OUTPUT_DIR / "chassis.json").read_text(encoding="utf-8"))


def infer_geometry_type(pieces: list[str]) -> str:
    """
    Déduit le type géométrique à partir des noms de pièces.
    Retourne un code de type géométrique.
    """
    pieces_lower = [p.lower() for p in pieces]
    has = lambda kw: any(kw in p for p in pieces_lower)
    count = lambda kw: sum(1 for p in pieces_lower if kw in p)

    nb = len(pieces)

    # Coulissants
    if has("vantail principal coulissant") or has("chicane") or has("percussion"):
        return "coulissant"
    if has("galandage"):
        return "galandage"

    # Portes (porte crémone, serrure, service) avec soubassement
    # → traverse intermédiaire sépare vitrage et panneau, soubassement hauteur fixe 0.45m
    if has("traverse intermédiaire") and not has("impos"):
        if has("meneau"):
            return "porte_capitales"   # CAPITALES avec meneaux
        return "porte_soubassement"    # Portes crémone/serrure/service standard

    # Fenêtre 1 vantail + imposte soufflet (traverse horizontale + meneau dormant)
    if has("impos") or (has("meneau dormant") and nb == 5 and not has("masse")):
        return "1_vantail_imposte"

    # Fenêtre 1 vantail + traverse intermédiaire (PRIMELIS spécial)
    if has("ti") and nb == 5:
        return "1_vantail_traverse_intermediaire"

    # Fenêtres à frappe standard
    if has("meneau dormant") and has("montant"):
        meneau_count = count("meneau")
        if meneau_count >= 3:
            return "2_vantaux_2_fixes_lateraux"
        if meneau_count == 2:
            if has("fixe"):
                return "2_vantaux_2_fixes_lateraux"
            return "4_vantaux"
        if has("fixe"):
            return "2_vantaux_1_fixe_lateral"
        if has("masse centrale"):
            return "3_vantaux"

    if has("masse centrale") and not has("meneau"):
        return "2_vantaux"

    if nb == 4:
        return "1_vantail"  # Fixe, 1 vantail, soufflet (même géométrie de cadre)

    return f"custom_{nb}pieces"


def extract_pieces(ws_val, ws_fml, sheet_name: str) -> dict | None:
    """
    Extrait les pièces (Af, Uf) et le type géométrique d'un onglet châssis.
    Retourne None si l'onglet est invalide.
    """
    pieces = []

    # Lire les pièces (lignes 9 à 20 max, col H=7, M=12, N=13)
    for row_val, row_fml in zip(
        ws_val.iter_rows(min_row=9, max_row=22, values_only=True),
        ws_fml.iter_rows(min_row=9, max_row=22, values_only=False),
    ):
        piece_name = row_val[7]  # col H
        af_val = row_val[12]     # col M
        uf_val = row_val[13]     # col N

        # La ligne suivante (juste après les pièces) contient "Ufi moyen"
        if piece_name == "Ufi moyen":
            break
        if piece_name is None or not isinstance(piece_name, str):
            continue
        if not isinstance(af_val, (int, float)) or not isinstance(uf_val, (int, float)):
            continue

        pieces.append({
            "nom": piece_name.strip(),
            "Af": round(float(af_val), 5),
            "Uf": round(float(uf_val), 4),
        })

    if not pieces:
        return None

    # Lire le Ufi moyen (ligne avec "Ufi moyen" dans col M)
    ufi_moyen = None
    for row_val in ws_val.iter_rows(min_row=9, max_row=22, values_only=True):
        if row_val[12] == "Ufi moyen" and isinstance(row_val[13], (int, float)):
            ufi_moyen = round(float(row_val[13]), 4)
            break

    # Lire les formules des zones vitrage (lignes 28-70, col B et C/E)
    formulas = {}
    for row in ws_fml.iter_rows(min_row=28, max_row=80, values_only=False):
        label = row[1].value   # col B
        val_c = row[2].value   # col C
        val_e = row[4].value   # col E (Uw result cell)

        if label is None:
            continue
        formula = val_e or val_c
        if formula and isinstance(formula, str) and formula.startswith("="):
            formulas[str(label).strip()] = formula

    # Déterminer si le châssis a un renfort (renfort = Oui dans ligne 18)
    renfort = "Non"
    for row_val in ws_val.iter_rows(min_row=16, max_row=22, values_only=True):
        if row_val[0] == "Renfort" or row_val[1] == "Renfort":
            r_val = row_val[2]
            if r_val in ("Oui", "Non"):
                renfort = r_val
            break

    piece_names = [p["nom"] for p in pieces]
    geo_type = infer_geometry_type(piece_names)

    # Paramètres spéciaux : soubassement (hauteur panneau fixe 0.45m pour portes)
    extra_params = {}
    if geo_type == "porte_soubassement":
        for formula in formulas.values():
            if "0.45" in formula:
                extra_params["hauteur_soubassement"] = 0.45
                break

    # Extraire les zones Psi_g/Ug par zone vitrage depuis les FORMULES (data_only=False)
    # Pour les valeurs hardcodées, lire la valeur de la cellule formula (pas data_only)
    zones_psi = []
    pending_ug: dict | None = None

    for rf in ws_fml.iter_rows(min_row=28, max_row=80, values_only=False):
        label_cell = rf[1]  # col B
        cell_c = rf[2]      # col C

        label = str(label_cell.value or "").strip().lower()
        cell_val = cell_c.value
        is_formula = isinstance(cell_val, str) and cell_val.startswith("=")
        is_hardcoded = isinstance(cell_val, (int, float))

        # Stop à la section Calcul Sw
        if label in ("calcul sw", "surface totale af+ag"):
            break

        # Ligne Ug (U-glass, U-pannel, U-panel, U-panneau)
        if any(label.startswith(k) for k in ("u-glass", "u-pannel", "u-panel", "u-panneau")):
            if is_formula:
                pending_ug = {"Ug_default": None, "Ug_dynamic": True}
            elif is_hardcoded:
                pending_ug = {"Ug_default": round(float(cell_val), 4), "Ug_dynamic": False}

        # Ligne PSI
        elif label.startswith("psi") and pending_ug is not None:
            if is_formula:
                zones_psi.append({
                    "zone": len(zones_psi),
                    **pending_ug,
                    "Psi_g_default": None,
                    "Psi_g_dynamic": True,
                })
            elif is_hardcoded:
                zones_psi.append({
                    "zone": len(zones_psi),
                    **pending_ug,
                    "Psi_g_default": round(float(cell_val), 5),
                    "Psi_g_dynamic": False,
                })
            pending_ug = None

    return {
        "pieces": pieces,
        "Ufi_moyen": ufi_moyen,
        "nb_pieces": len(pieces),
        "geo_type": geo_type,
        "renfort": renfort,
        "extra_params": extra_params,
        "zones_psi": zones_psi,
    }


def main():
    geometry_data = {}
    errors = []

    # Récupérer tous les IDs d'onglets valides
    sheet_ids = set()
    for c in chassis_list:
        sheet_ids.add(str(c["id"]))

    valid_sheets = [s for s in wb_val.sheetnames if s in sheet_ids]
    print(f"Traitement de {len(valid_sheets)} onglets châssis...")

    for sheet_name in valid_sheets:
        try:
            ws_val = wb_val[sheet_name]
            ws_fml = wb_fml[sheet_name]
            data = extract_pieces(ws_val, ws_fml, sheet_name)
            if data:
                geometry_data[int(sheet_name)] = data
            else:
                errors.append(f"Onglet {sheet_name}: aucune pièce trouvée")
        except Exception as e:
            errors.append(f"Onglet {sheet_name}: {e}")

    # Statistiques par type géométrique
    geo_types_count: dict[str, int] = {}
    for data in geometry_data.values():
        t = data["geo_type"]
        geo_types_count[t] = geo_types_count.get(t, 0) + 1

    print(f"\n✓ {len(geometry_data)} onglets extraits avec succès")
    print(f"✗ {len(errors)} erreurs")
    print("\nTypes géométriques détectés :")
    for t, count in sorted(geo_types_count.items(), key=lambda x: -x[1]):
        print(f"  {t:40s} : {count}")

    if errors:
        print("\nErreurs :")
        for e in errors[:20]:
            print(f"  {e}")

    output_path = OUTPUT_DIR / "chassis_geometry.json"
    output_path.write_text(
        json.dumps(geometry_data, ensure_ascii=False, indent=2, default=str),
        encoding="utf-8",
    )
    print(f"\nÉcrit : {output_path}")


if __name__ == "__main__":
    main()
